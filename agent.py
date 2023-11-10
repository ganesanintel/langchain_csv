from langchain import OpenAI
from langchain.agents import create_pandas_dataframe_agent
import tabula
from langchain.chat_models import AzureChatOpenAI
from io import BytesIO
from pytesseract import image_to_string
from PIL import Image
import pandas as pd
# pdftotree is one way to parse complex tables. Have to find the content with tables and then take them through this.. hmm, more logistical overhead!
import pdftotree
import pypdfium2 as pdfium
# another exploration
import camelot
from img2table.document import PDF
from img2table.ocr import TesseractOCR
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
import csv

# Setting up the api key
import environ

env = environ.Env()
environ.Env.read_env()

API_KEY = env("apikey")


def create_agent(filename: str):
    """
    Create an agent that can access and use a large language model (LLM).

    Args:
        filename: The path to the CSV file that contains the data.

    Returns:
        An agent that can access and use the LLM.
    """

    # Create an OpenAI object.
#    llm = OpenAI(openai_api_key=API_KEY)
    llm=AzureChatOpenAI(
            deployment_name="GPT_45",
            openai_api_base="https://nex-mi6-gpt4-production.openai.azure.com",
            openai_api_version="2023-07-01-preview",
            openai_api_key=API_KEY,
            openai_api_type="azure",
            streaming=False,
            temperature=0,
        )

    print("------")
    print(filename)
    print(type(filename))
    print("------")
    if(filename.type == "text/csv"):
        # Read the CSV file into a Pandas DataFrame.
        df = pd.read_csv(filename)
        return create_pandas_dataframe_agent(llm, df, verbose=False)
    elif(filename.type == "application/pdf"):
        file="./" + filename.name

        print(file)
        with open(file, "wb") as f:
            f.write(filename.getvalue())
        
        pdf = PDF(src=file)
        ocr = TesseractOCR(lang="eng")
        pdf_tables = pdf.extract_tables(ocr=ocr)
        pdf.to_xlsx('tables.xlsx',ocr=ocr)

        wb = load_workbook(filename = 'tables.xlsx')

        for st_name in wb.sheetnames:
            st = wb[st_name]
            mcr_coord_list = [mcr.coord for mcr in st.merged_cells.ranges]
            print(mcr_coord_list)

            for mcr in mcr_coord_list:
                min_col, min_row, max_col, max_row = range_boundaries(mcr)
                top_left_cell_value = st.cell(row=min_row, column=min_col).value
                st.unmerge_cells(mcr)
                for row in st.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                    for cell in row:
                        cell.value = top_left_cell_value

        wb.save('merged_table.xlsx')

        wb = load_workbook(filename='merged_table.xlsx')

        sheetnames = wb.get_sheet_names()
        sheet_value_arr = []
        df_arr = []
        dummyIndex = 0
        for a in sheetnames:
            sh = wb[a]
            filename = a.replace(" ", "-")
            with open(filename +'.csv', 'w') as file_handle:
                csv_writer = csv.writer(file_handle)
                for row in sh.iter_rows():
                    tmp = []
                    for cell in row:
                        if cell.value:
                            tmp.append(cell.value.replace("\n", " "))
                        else:
                            tmp.append(" ")
                    #csv_writer.writerow([cell.value.replace("\n", " ") for cell in row])
                    csv_writer.writerow(tmp) 
            try:
                """df = pd.read_csv(filename +'.csv')
                newDict = {}
                newDict['table_data'] = df.values.tolist()
                if dummyIndex == 0:
                    newDict['content'] = "State census table by 2005"
                else:
                    newDict['content'] = "State census table by 2010"

                df_arr.append(pd.DataFrame.from_dict(newDict))"""
                df_arr.append(pd.read_csv(filename +'.csv'))
            except:
                print("Error in "+filename+'.csv')
                print(pd.errors)
                pass

            dummyIndex = dummyIndex + 1

        #df = pd.read_csv(filename)
        print("Creating data frames")
        return create_pandas_dataframe_agent(llm, df_arr, verbose=False)

def convert_pdf_to_images(file_path, scale=300/72):

    pdf_file = pdfium.PdfDocument(file_path)

    page_indices = [i for i in range(len(pdf_file))]

    renderer = pdf_file.render(
        pdfium.PdfBitmap.to_pil,
        page_indices=page_indices,
        scale=scale,
    )

    final_images = []

    for i, image in zip(page_indices, renderer):

        image_byte_array = BytesIO()
        image.save(image_byte_array, format='jpeg', optimize=True)
        image_byte_array = image_byte_array.getvalue()
        final_images.append(dict({i: image_byte_array}))

    return final_images

# 2. Extract text from images via pytesseract


def extract_text_from_img(list_dict_final_images):

    image_list = [list(data.values())[0] for data in list_dict_final_images]
    image_content = []

    for index, image_bytes in enumerate(image_list):

        image = Image.open(BytesIO(image_bytes))
        raw_text = str(image_to_string(image))
        image_content.append(raw_text)

    return "\n".join(image_content)



def query_agent(agent, query):
    """
    Query an agent and return the response as a string.

    Args:
        agent: The agent to query.
        query: The query to ask the agent.

    Returns:
        The response from the agent as a string.
    """

    prompt = (
        """
            For the following query, if it requires drawing a table, reply as follows:
            {"table": {"columns": ["column1", "column2", ...], "data": [[value1, value2, ...], [value1, value2, ...], ...]}}

            If the query requires creating a bar chart, reply as follows:
            {"bar": {"columns": ["A", "B", "C", ...], "data": [25, 24, 10, ...]}}
            
            If the query requires creating a line chart, reply as follows:
            {"line": {"columns": ["A", "B", "C", ...], "data": [25, 24, 10, ...]}}
            
            There can only be two types of chart, "bar" and "line".
            
            If it is just asking a question that requires neither, reply as follows:
            {"answer": "answer"}
            Example:
            {"answer": "The title with the highest rating is 'Gilead'"}
            
            If you do not know the answer, reply as follows:
            {"answer": "I do not know."}
            
            Return all output as a string.
            
            All strings in "columns" list and data list, should be in double quotes,
            
            For example: {"columns": ["title", "ratings_count"], "data": [["Gilead", 361], ["Spider's Web", 5164]]}
            
            Lets think step by step.
            
            Below is the query.
            Query: 
            """
        + query
    )

    # Run the prompt through the agent.
    response = agent.run(prompt)

    # Convert the response to a string.
    return response.__str__()
